"""Export API endpoints for device data"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import datetime, timedelta, timezone
from typing import Optional
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.database import get_db
from app.models.models import Device, DeviceReading
from app.core.logging_config import get_logger

logger = get_logger("export_api")
router = APIRouter()

# Pakistan Standard Time (PKT) is UTC+5
PKT = timezone(timedelta(hours=5))


def convert_to_pkt(dt: datetime) -> datetime:
    """Convert datetime to Pakistan Standard Time (UTC+5)"""
    if dt is None:
        return None
    # If datetime is naive (no timezone), assume it's UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(PKT)


def create_csv_export(readings: list, device_info: dict) -> str:
    """Create CSV export from device readings"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Device ID',
        'Device Name',
        'Timestamp',
        'Temperature (°F)',
        'Static Pressure (PSI)',
        'Differential Pressure (IWC)',
        'Volume (MCF)',
        'Total Volume Flow (MCF/day)',
        'Battery (%)'
    ])

    # Write data rows
    for reading in readings:
        # Convert timestamp to Pakistan Standard Time (UTC+5)
        pkt_timestamp = convert_to_pkt(reading.timestamp)
        writer.writerow([
            device_info.get('client_id', ''),
            device_info.get('device_name', ''),
            pkt_timestamp.strftime('%Y-%m-%d %H:%M:%S') if pkt_timestamp else '',
            f"{reading.temperature:.2f}",
            f"{reading.static_pressure:.2f}",
            f"{reading.differential_pressure:.2f}",
            f"{reading.volume:.2f}",
            f"{reading.total_volume_flow:.2f}",
            f"{(reading.battery or 0):.0f}"
        ])

    return output.getvalue()


def create_excel_export(readings: list, device_info: dict) -> bytes:
    """Create Excel export from device readings"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Device Readings"

    # Style for header
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header
    headers = [
        'Device ID',
        'Device Name',
        'Timestamp',
        'Temperature (°F)',
        'Static Pressure (PSI)',
        'Differential Pressure (IWC)',
        'Volume (MCF)',
        'Total Volume Flow (MCF/day)',
        'Battery (%)'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, reading in enumerate(readings, 2):
        # Convert timestamp to Pakistan Standard Time (UTC+5)
        pkt_timestamp = convert_to_pkt(reading.timestamp)
        ws.cell(row=row_num, column=1, value=device_info.get('client_id', ''))
        ws.cell(row=row_num, column=2, value=device_info.get('device_name', ''))
        ws.cell(row=row_num, column=3, value=pkt_timestamp.strftime('%Y-%m-%d %H:%M:%S') if pkt_timestamp else '')
        ws.cell(row=row_num, column=4, value=f"{reading.temperature:.2f}")
        ws.cell(row=row_num, column=5, value=f"{reading.static_pressure:.2f}")
        ws.cell(row=row_num, column=6, value=f"{reading.differential_pressure:.2f}")
        ws.cell(row=row_num, column=7, value=f"{reading.volume:.2f}")
        ws.cell(row=row_num, column=8, value=f"{reading.total_volume_flow:.2f}")
        ws.cell(row=row_num, column=9, value=f"{(reading.battery or 0):.0f}")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_multi_device_csv_export(devices_data: list) -> str:
    """Create CSV export for multiple devices"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Device ID',
        'Device Name',
        'Section',
        'Location',
        'Timestamp',
        'Temperature (°F)',
        'Static Pressure (PSI)',
        'Differential Pressure (IWC)',
        'Volume (MCF)',
        'Total Volume Flow (MCF/day)',
        'Battery (%)'
    ])

    # Write data rows for all devices
    for device_data in devices_data:
        device = device_data['device']
        for reading in device_data['readings']:
            # Convert timestamp to Pakistan Standard Time (UTC+5)
            pkt_timestamp = convert_to_pkt(reading.timestamp)
            writer.writerow([
                device.client_id,
                device.device_name,
                device.client_id.split('-')[0] if '-' in device.client_id else 'N/A',
                device.location,
                pkt_timestamp.strftime('%Y-%m-%d %H:%M:%S') if pkt_timestamp else '',
                f"{reading.temperature:.2f}",
                f"{reading.static_pressure:.2f}",
                f"{reading.differential_pressure:.2f}",
                f"{reading.volume:.2f}",
                f"{reading.total_volume_flow:.2f}",
                f"{(reading.battery or 0):.0f}"
            ])

    return output.getvalue()


def create_multi_device_excel_export(devices_data: list) -> bytes:
    """Create Excel export for multiple devices"""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Devices Readings"

    # Style for header
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header
    headers = [
        'Device ID',
        'Device Name',
        'Section',
        'Location',
        'Timestamp',
        'Temperature (°F)',
        'Static Pressure (PSI)',
        'Differential Pressure (IWC)',
        'Volume (MCF)',
        'Total Volume Flow (MCF/day)',
        'Battery (%)'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows for all devices
    row_num = 2
    for device_data in devices_data:
        device = device_data['device']
        for reading in device_data['readings']:
            # Convert timestamp to Pakistan Standard Time (UTC+5)
            pkt_timestamp = convert_to_pkt(reading.timestamp)
            ws.cell(row=row_num, column=1, value=device.client_id)
            ws.cell(row=row_num, column=2, value=device.device_name)
            ws.cell(row=row_num, column=3, value=device.client_id.split('-')[0] if '-' in device.client_id else 'N/A')
            ws.cell(row=row_num, column=4, value=device.location)
            ws.cell(row=row_num, column=5, value=pkt_timestamp.strftime('%Y-%m-%d %H:%M:%S') if pkt_timestamp else '')
            ws.cell(row=row_num, column=6, value=f"{reading.temperature:.2f}")
            ws.cell(row=row_num, column=7, value=f"{reading.static_pressure:.2f}")
            ws.cell(row=row_num, column=8, value=f"{reading.differential_pressure:.2f}")
            ws.cell(row=row_num, column=9, value=f"{reading.volume:.2f}")
            ws.cell(row=row_num, column=10, value=f"{reading.total_volume_flow:.2f}")
            ws.cell(row=row_num, column=11, value=f"{(reading.battery or 0):.0f}")
            row_num += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/device/{device_id}")
async def export_device_data(
    device_id: str,
    start: str = Query(..., description="Start datetime in ISO format"),
    end: str = Query(..., description="End datetime in ISO format"),
    format: str = Query("csv", description="Export format: csv or excel"),
    db: Session = Depends(get_db)
):
    """Export data for a single device"""
    try:
        # Parse dates
        start_date = datetime.fromisoformat(start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end.replace('Z', '+00:00'))

        # Get device - try by client_id first, then by database ID
        device = db.query(Device).filter(Device.client_id == device_id).first()
        if not device and device_id.isdigit():
            # Try by database ID if device_id is numeric
            device = db.query(Device).filter(Device.id == int(device_id)).first()
        if not device:
            raise HTTPException(status_code=404, detail="Device not found")

        # Get readings
        readings = db.query(DeviceReading).filter(
            and_(
                DeviceReading.device_id == device.id,
                DeviceReading.timestamp >= start_date,
                DeviceReading.timestamp <= end_date
            )
        ).order_by(DeviceReading.timestamp.desc()).all()

        if not readings:
            raise HTTPException(status_code=404, detail="No readings found for the specified time range")

        device_info = {
            'client_id': device.client_id,
            'device_name': device.device_name
        }

        # Generate export based on format
        if format == "csv":
            content = create_csv_export(readings, device_info)
            filename = f"device_{device_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
            media_type = "text/csv"
        elif format == "excel":
            content = create_excel_export(readings, device_info)
            filename = f"device_{device_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")

        logger.info(f"[EXPORT] Generated {format.upper()} export for device {device_id}: {len(readings)} readings")

        return StreamingResponse(
            io.BytesIO(content.encode() if isinstance(content, str) else content),
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"[ERROR] Export failed for device {device_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/section/{section_id}")
async def export_section_data(
    section_id: str,
    start: str = Query(..., description="Start datetime in ISO format"),
    end: str = Query(..., description="End datetime in ISO format"),
    format: str = Query("csv", description="Export format: csv or excel"),
    db: Session = Depends(get_db)
):
    """Export data for all devices in a section"""
    try:
        # Parse dates
        start_date = datetime.fromisoformat(start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end.replace('Z', '+00:00'))

        # Get all devices in section
        if section_id == "OTHER":
            # For OTHER section, get devices with device_type = 'OTHER'
            devices = db.query(Device).filter(Device.device_type == "OTHER").all()
        else:
            # For SMS sections (I, II, III, IV, V), support both SMS-I-XXX and SMS-1-XXX formats
            section_map = {'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5'}
            arabic_num = section_map.get(section_id, section_id)

            devices = db.query(Device).filter(
                or_(
                    Device.client_id.like(f"SMS-{section_id}-%"),
                    Device.client_id.like(f"SMS-{arabic_num}-%")
                )
            ).all()

        if not devices:
            raise HTTPException(status_code=404, detail=f"No devices found in section {section_id}")

        # Collect data for all devices
        devices_data = []
        total_readings = 0

        for device in devices:
            readings = db.query(DeviceReading).filter(
                and_(
                    DeviceReading.device_id == device.id,
                    DeviceReading.timestamp >= start_date,
                    DeviceReading.timestamp <= end_date
                )
            ).order_by(DeviceReading.timestamp.desc()).all()

            if readings:
                devices_data.append({
                    'device': device,
                    'readings': readings
                })
                total_readings += len(readings)

        if not devices_data:
            raise HTTPException(status_code=404, detail="No readings found for the specified time range")

        # Generate export based on format
        if format == "csv":
            content = create_multi_device_csv_export(devices_data)
            filename = f"section_{section_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
            media_type = "text/csv"
        elif format == "excel":
            content = create_multi_device_excel_export(devices_data)
            filename = f"section_{section_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")

        logger.info(f"[EXPORT] Generated {format.upper()} export for section {section_id}: {len(devices_data)} devices, {total_readings} readings")

        return StreamingResponse(
            io.BytesIO(content.encode() if isinstance(content, str) else content),
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"[ERROR] Export failed for section {section_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/all")
async def export_all_devices_data(
    start: str = Query(..., description="Start datetime in ISO format"),
    end: str = Query(..., description="End datetime in ISO format"),
    format: str = Query("csv", description="Export format: csv or excel"),
    db: Session = Depends(get_db)
):
    """Export data for all SMS devices across all sections (I-V)"""
    try:
        # Parse dates
        start_date = datetime.fromisoformat(start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end.replace('Z', '+00:00'))

        # Get all SMS devices (sections I-V)
        devices = db.query(Device).filter(
            Device.device_type == "SMS"
        ).all()

        if not devices:
            raise HTTPException(status_code=404, detail="No SMS devices found")

        # Collect data for all devices
        devices_data = []
        total_readings = 0

        for device in devices:
            readings = db.query(DeviceReading).filter(
                and_(
                    DeviceReading.device_id == device.id,
                    DeviceReading.timestamp >= start_date,
                    DeviceReading.timestamp <= end_date
                )
            ).order_by(DeviceReading.timestamp.desc()).all()

            if readings:
                devices_data.append({
                    'device': device,
                    'readings': readings
                })
                total_readings += len(readings)

        if not devices_data:
            raise HTTPException(status_code=404, detail="No readings found for the specified time range")

        # Generate export based on format
        if format == "csv":
            content = create_multi_device_csv_export(devices_data)
            filename = f"all_devices_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
            media_type = "text/csv"
        elif format == "excel":
            content = create_multi_device_excel_export(devices_data)
            filename = f"all_devices_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")

        logger.info(f"[EXPORT] Generated {format.upper()} export for all devices: {len(devices_data)} devices, {total_readings} readings")

        return StreamingResponse(
            io.BytesIO(content.encode() if isinstance(content, str) else content),
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"[ERROR] Export failed for all devices: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
