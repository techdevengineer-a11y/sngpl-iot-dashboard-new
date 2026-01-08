"""Export service for generating PDF and Excel reports"""

from datetime import datetime
from io import BytesIO
from typing import List, Optional
from sqlalchemy.orm import Session

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.models import Device, DeviceReading, Alarm
from app.core.logging_config import get_logger

logger = get_logger(__name__)


class ExportService:
    """Service for exporting data to PDF and Excel formats"""

    @staticmethod
    def generate_devices_pdf(devices: List[Device], db: Session) -> BytesIO:
        """Generate PDF report for devices"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("SNGPL IoT Platform - Devices Report", title_style))
        elements.append(Spacer(1, 12))

        # Report metadata
        meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        elements.append(Paragraph(f"Total Devices: {len(devices)}", meta_style))
        elements.append(Spacer(1, 20))

        # Table data
        table_data = [['Device ID', 'Client ID', 'Location', 'Type', 'Status', 'Last Seen']]

        for device in devices:
            status = "Active" if device.is_active else "Inactive"
            last_seen = device.last_seen.strftime('%Y-%m-%d %H:%M') if device.last_seen else "Never"
            table_data.append([
                str(device.id),
                device.client_id,
                device.location or "Unknown",
                device.device_type or "N/A",
                status,
                last_seen
            ])

        # Create table
        table = Table(table_data, colWidths=[0.8*inch, 1.2*inch, 1.5*inch, 0.8*inch, 0.8*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_devices_excel(devices: List[Device], db: Session) -> BytesIO:
        """Generate Excel report for devices"""
        buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Devices Report"

        # Header styling
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = "SNGPL IoT Platform - Devices Report"
        title_cell.font = Font(bold=True, size=16, color="1e40af")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Metadata
        sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A3'] = f"Total Devices: {len(devices)}"

        # Headers
        headers = ['Device ID', 'Client ID', 'Location', 'Type', 'Status', 'Last Seen']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        for row_num, device in enumerate(devices, 6):
            status = "Active" if device.is_active else "Inactive"
            last_seen = device.last_seen.strftime('%Y-%m-%d %H:%M') if device.last_seen else "Never"

            row_data = [
                device.id,
                device.client_id,
                device.location or "Unknown",
                device.device_type or "N/A",
                status,
                last_seen
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Color code status
                if col_num == 5:  # Status column
                    if status == "Active":
                        cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

        # Adjust column widths
        for col_num in range(1, 7):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 18

        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_alarms_pdf(alarms: List[Alarm], db: Session) -> BytesIO:
        """Generate PDF report for alarms"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#dc2626'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("SNGPL IoT Platform - Alarms Report", title_style))
        elements.append(Spacer(1, 12))

        # Report metadata
        meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        elements.append(Paragraph(f"Total Alarms: {len(alarms)}", meta_style))
        elements.append(Spacer(1, 20))

        # Table data
        table_data = [['Device', 'Parameter', 'Value', 'Severity', 'Type', 'Status', 'Time']]

        for alarm in alarms:
            status = "Resolved" if alarm.is_acknowledged else "Active"
            created = alarm.triggered_at.strftime('%m-%d %H:%M') if alarm.triggered_at else "N/A"
            table_data.append([
                alarm.client_id[:12],
                alarm.parameter[:15],
                f"{alarm.value:.2f}",
                alarm.severity.upper(),
                alarm.threshold_type,
                status,
                created
            ])

        # Create table
        table = Table(table_data, colWidths=[1.1*inch, 1.1*inch, 0.8*inch, 0.9*inch, 0.8*inch, 0.9*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_alarms_excel(alarms: List[Alarm], db: Session) -> BytesIO:
        """Generate Excel report for alarms"""
        buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Alarms Report"

        # Header styling
        header_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        sheet.merge_cells('A1:G1')
        title_cell = sheet['A1']
        title_cell.value = "SNGPL IoT Platform - Alarms Report"
        title_cell.font = Font(bold=True, size=16, color="dc2626")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Metadata
        sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A3'] = f"Total Alarms: {len(alarms)}"

        # Headers
        headers = ['Device', 'Parameter', 'Value', 'Severity', 'Type', 'Status', 'Time']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        for row_num, alarm in enumerate(alarms, 6):
            status = "Resolved" if alarm.is_acknowledged else "Active"
            created = alarm.triggered_at.strftime('%Y-%m-%d %H:%M') if alarm.triggered_at else "N/A"

            row_data = [
                alarm.client_id,
                alarm.parameter,
                round(alarm.value, 2),
                alarm.severity.upper(),
                alarm.threshold_type,
                status,
                created
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Color code severity
                if col_num == 4:  # Severity column
                    if alarm.severity == "high":
                        cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                    elif alarm.severity == "medium":
                        cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")

        # Adjust column widths
        for col_num in range(1, 8):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 16

        workbook.save(buffer)
        buffer.seek(0)
        return buffer


# Global instance
export_service = ExportService()
